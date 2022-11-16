import json
import logging
import datetime as dt
import numpy as np
from functools import cached_property
from gcm.Dao.DaoRunner import DaoRunner, DaoRunnerConfigArgs
import pandas as pd
from gcm.Dao.DaoSources import DaoSource
from gcm.inv.dataprovider.strategy_benchmark import StrategyBenchmark
from gcm.inv.dataprovider.investment_group import InvestmentGroup
from gcm.inv.dataprovider.entity_master import EntityMaster
from gcm.Dao.daos.azure_datalake.azure_datalake_dao import AzureDataLakeDao
from pandas._libs.tslibs.offsets import relativedelta
from gcm.inv.quantlib.enum_source import Periodicity, PeriodicROR

from _legacy.Reports.reports.peer_rankings.peer_rankings import PeerRankings
from _legacy.core.ReportStructure.report_structure import (
    ReportingEntityTypes,
    ReportType,
    AggregateInterval, ReportVertical,
)
from _legacy.core.Runners.investmentsreporting import (
    InvestmentsReportRunner,
)
from gcm.inv.scenario import Scenario
from gcm.inv.quantlib.timeseries.analytics import Analytics
from _legacy.core.reporting_runner_base import (
    ReportingRunnerBase,
)
from gcm.inv.dataprovider.factor import Factor
from gcm.inv.quantlib.timeseries.transformer.aggregate_from_daily import (
    AggregateFromDaily,
)
from gcm.inv.quantlib.timeseries.transformer.beta import Beta
from openpyxl.utils.cell import get_column_letter


class PerformanceScreenerReport(ReportingRunnerBase):
    def __init__(self, peer_group, trailing_months=36):
        super().__init__(runner=Scenario.get_attribute("runner"))
        self._as_of_date = Scenario.get_attribute("as_of_date")
        self._end_date = self._as_of_date
        self._trailing_months = trailing_months
        self._start_date = self._as_of_date - relativedelta(years=self._trailing_months / 12) + relativedelta(days=1)
        self._analytics = Analytics()
        self._strategy_benchmark = StrategyBenchmark()
        self._peer_group = peer_group
        self._summary_data_location = "raw/investmentsreporting/summarydata/ars_performance_screener"

    @cached_property
    def _constituents(self):
        return self._get_peer_constituents()

    @cached_property
    def _all_constituent_returns(self):
        inv_group_ids = self._constituents['InvestmentGroupId'].unique().tolist()
        returns = self._get_constituent_returns(investment_group_ids=inv_group_ids)
        return returns

    @cached_property
    def _updated_constituent_returns(self, min_required_returns=24):
        as_of_month = dt.date(year=self._as_of_date.year, month=self._as_of_date.month, day=1)
        as_of_month = pd.to_datetime(as_of_month)

        if as_of_month not in self._all_constituent_returns.index:
            return None

        updated_funds = self._all_constituent_returns.loc[as_of_month].dropna().index
        returns = self._all_constituent_returns.loc[:, updated_funds]
        sufficient_track = ~returns[-min_required_returns:].isna().any()
        sufficient_track_funds = sufficient_track[sufficient_track].index.tolist()
        returns = returns.loc[:, sufficient_track_funds]

        not_all_zeros = ~(returns == 0).all()
        returns = returns.loc[:, not_all_zeros]

        return returns

    @cached_property
    def _absolute_benchmark_returns(self):
        ids = self._constituents['InvestmentGroupId'].unique().tolist()
        inv_group = InvestmentGroup(investment_group_ids=ids)
        abs_bmrk_returns = inv_group.get_absolute_benchmark_returns(start_date=self._start_date,
                                                                    end_date=self._end_date)
        abs_bmrk_returns = AggregateFromDaily().transform(
            data=abs_bmrk_returns,
            method="geometric",
            period=Periodicity.Monthly,
            first_of_day=True
        )
        group_dimensions = inv_group.get_dimensions()
        id_name_map = dict(zip(group_dimensions['InvestmentGroupId'], group_dimensions['InvestmentGroupName']))
        abs_bmrk_returns.columns = [id_name_map.get(item, item) for item in abs_bmrk_returns.columns]
        return abs_bmrk_returns

    @cached_property
    def _rf_returns(self):
        returns = self._get_monthly_factor_returns(ticker="SBMMTB1 Index")
        return returns

    @cached_property
    def _spx_returns(self):
        returns = self._get_monthly_factor_returns(ticker="SPX Index")
        return returns

    @cached_property
    def _peer_benchmark_return(self):
        returns = self._get_monthly_factor_returns(ticker=self._peer_benchmark_ticker)
        return returns

    def _get_json_file_name(self, as_of_date):
        as_of_date = as_of_date.strftime("%Y-%m-%d")
        file_name = self._peer_group.replace("/", "") + as_of_date + ".json"
        return file_name

    def _download_prior_rankings(self, location, as_of_date) -> dict:
        file_path = self._get_json_file_name(as_of_date=as_of_date)
        read_params = AzureDataLakeDao.create_get_data_params(
            location,
            file_path,
            retry=False,
        )
        try:
            file = self._runner.execute(
                params=read_params,
                source=DaoSource.DataLake,
                operation=lambda dao, params: dao.get_data(read_params),
            )
        except:
            return None

        inputs = json.loads(file.content)
        inputs = pd.read_json(inputs['summary_table'], orient="index")[['InvestmentGroupName', 'Decile']]
        inputs = inputs.iloc[inputs['Decile'].values != '', ]
        inputs.rename(columns={'Decile': as_of_date}, inplace=True)
        return inputs

    def _validate_inputs(self):
        pass

    def _get_altsoft_investment_names(self, entities):
        alt_soft_entities = entities[entities['SourceName'].str.contains('AltSoft.')]
        alt_soft_entities = alt_soft_entities[alt_soft_entities['IsReportingInvestment']]
        alt_soft_entities = alt_soft_entities[['InvestmentName', 'InvestmentGroupName']]
        return alt_soft_entities

    def _get_peer_constituents(self):
        constituents = self._strategy_benchmark.get_altsoft_peer_constituents(peer_names=self._peer_group)
        ids = constituents['InvestmentGroupId']
        investment_group = InvestmentGroup(investment_group_ids=ids)
        dimensions = investment_group.get_dimensions()
        dimensions = dimensions[['InvestmentGroupId', 'InvestmentGroupName', 'InvestmentStatus']]

        constituents = constituents.merge(dimensions, on=['InvestmentGroupId'], how='inner')

        entities = EntityMaster().get_investment_entities()
        reporting_inv_names = self._get_altsoft_investment_names(entities)
        constituents = constituents.merge(reporting_inv_names, on=['InvestmentGroupName'], how='left')

        constituents = constituents[['InvestmentGroupId', 'InvestmentGroupName', 'InvestmentName', 'InvestmentStatus']]
        constituents = constituents.drop_duplicates()

        return constituents

    def _get_constituent_returns(self, investment_group_ids):
        investment_group = InvestmentGroup(investment_group_ids=investment_group_ids)
        returns = investment_group.get_monthly_returns(start_date=self._start_date,
                                                       end_date=self._as_of_date,
                                                       wide=True)
        return returns

    def _get_monthly_factor_returns(self, ticker):
        factors = Factor(tickers=[ticker])
        returns = factors.get_returns(
            start_date=self._start_date,
            end_date=self._end_date,
            fill_na=True,
        )
        if len(returns) > 0:
            returns = AggregateFromDaily().transform(
                data=returns,
                method="geometric",
                period=Periodicity.Monthly,
            )
            returns.index = [dt.datetime(x.year, x.month, 1) for x in returns.index.tolist()]
            return returns
        else:
            return pd.DataFrame()

    def _calculate_annualized_return(self, returns):
        ann_factor = np.minimum(1, 12 / returns.notna().sum())
        cumulative_returns = self._analytics.compute_periodic_return(
            ror=returns,
            period=PeriodicROR.ITD,
            as_of_date=self._as_of_date,
            method="geometric",
        )

        annualized_returns = (1 + cumulative_returns) ** (ann_factor) - 1

        ror = annualized_returns.to_frame('Return')
        return ror

    def _get_annualized_return(self):
        returns = self._calculate_annualized_return(returns=self._updated_constituent_returns)
        return returns

    def _get_annualized_vol(self):
        # vol = self._analytics.compute_trailing_vol(
        #     ror=self._constituent_returns,
        #     window=self._trailing_months,
        #     as_of_date=self._as_of_date,
        #     periodicity=Periodicity.Monthly,
        #     annualize=True,
        # )
        vol = self._updated_constituent_returns.std() * np.sqrt(12)
        vol = vol.to_frame('Vol')
        return vol

    def _get_sharpe_ratio(self):
        # sharpe = self._analytics.compute_trailing_sharpe_ratio(
        #     ror=self._updated_constituent_returns,
        #     rf_ror=self._rf_returns,
        #     window=self._trailing_months,
        #     as_of_date=self._as_of_date,
        #     periodicity=Periodicity.Monthly,
        # )
        #
        rf = (self._rf_returns.mean() * 12).squeeze()
        excess_return = (self._get_annualized_return() - rf)
        vol = self._get_annualized_vol()
        sharpe = excess_return['Return'] / vol['Vol']
        sharpe = sharpe.to_frame('Sharpe')
        return sharpe

    def _get_max_drawdown(self):
        wealth_index = 1000 * (1 + self._updated_constituent_returns).cumprod()
        previous_peaks = wealth_index.cummax()
        drawdowns = (wealth_index - previous_peaks) / previous_peaks
        max_drawdown = drawdowns.min()
        max_drawdown = max_drawdown.to_frame('MaxPttDdown')
        return max_drawdown

    def _get_return_in_peer_stress_months(self):
        avg_peer_return = self._updated_constituent_returns.median(axis=1)
        number_months = round(self._trailing_months * 0.10)
        worst_peer_months = avg_peer_return.sort_values()[0:number_months]
        worst_peer_dates = worst_peer_months.index.tolist()
        peer_stress_ror = self._updated_constituent_returns.loc[worst_peer_dates].mean(axis=0)
        peer_stress_ror = peer_stress_ror.to_frame('PeerStressRor')
        return peer_stress_ror

    def _get_beta_to_spx(self):
        trailing_betas = Beta().transform(data=self._updated_constituent_returns,
                                          benchmark=self._spx_returns,
                                          period=Periodicity.Monthly,
                                          window=36)
        trailing_betas = trailing_betas.iloc[-1, :]
        trailing_betas = trailing_betas.to_frame('BetaToSpx')
        return trailing_betas

    def _get_max_1mo_return(self):
        max_ror = self._updated_constituent_returns.max(axis=0)
        max_ror = max_ror.to_frame('MaxRor')
        return max_ror

    def _get_calendar_returns(self):
        returns = self._updated_constituent_returns.copy()
        returns = returns + 1
        cumulative_returns = returns.groupby(returns.index.year).prod()
        returns = cumulative_returns - 1
        returns = returns.T
        returns.rename(columns={returns.columns[0]: str(returns.columns[0]) + ' (Partial)'}, inplace=True)
        returns.rename(columns={returns.columns[-1]: str(returns.columns[-1]) + ' YTD'}, inplace=True)
        returns = returns.iloc[:, ::-1]

        if returns.shape[1] == 3:
            returns = pd.concat([returns, pd.DataFrame(columns=['-'], index=returns.index)], axis=1)

        return returns

    def _calculate_peer_rankings(self):
        rankings = PeerRankings().calculate_peer_rankings(peer_group_name=self._peer_group,
                                                          peer_group_returns=self._updated_constituent_returns)
        rankings.rename(columns={'rank': 'Rank', 'InvestmentGroupId': 'InvestmentGroupName',
                                 'confidence': 'Confidence'}, inplace=True)
        inv_names = self._constituents[['InvestmentGroupName', 'InvestmentGroupId']].drop_duplicates()
        rankings = rankings.merge(inv_names, how='left')
        rankings = rankings[['InvestmentGroupId', 'InvestmentGroupName', 'Rank', 'Confidence']]

        rankings['Confidence'] = rankings['Confidence'].astype(str) + '/5'
        rankings['Quartile'] = pd.qcut(x=rankings['Rank'], q=[x / 100 for x in range(0, 110, 25)],
                                       labels=[x for x in range(1, 5)])
        rankings['Decile'] = pd.qcut(x=rankings['Rank'], q=[x / 100 for x in range(0, 110, 10)],
                                     labels=[x / 10 for x in range(10, 110, 10)])
        # rankings['Rank'] = rankings['Points'].rank(pct=False, ascending=False)
        rankings = rankings.sort_values(['Rank', 'InvestmentGroupName'], ascending=[True, True])
        rankings = rankings[['Rank', 'Decile', 'Confidence', 'InvestmentGroupName', 'Quartile']]

        return rankings

    def get_header_info(self):
        header = pd.DataFrame(
            {
                "header_info": [
                    self._peer_group,
                    'ARS',
                    self._as_of_date,
                ]
            }
        )
        return header

    def build_constituent_count_summary(self):
        counts = [self._updated_constituent_returns.shape[1], self._constituents.shape[0]]

        counts = pd.DataFrame({'counts': counts})
        return counts

    def build_standalone_metrics_summary(self):
        ror = self._get_annualized_return()
        vol = self._get_annualized_vol()
        sharpe = self._get_sharpe_ratio()
        max_ptt_ddown = self._get_max_drawdown()
        beta = self._get_beta_to_spx()
        calendar_returns = self._get_calendar_returns()

        summary = ror.merge(vol, how='left', left_index=True, right_index=True)
        summary = summary.merge(sharpe, how='left', left_index=True, right_index=True)
        summary = summary.merge(max_ptt_ddown, how='left', left_index=True, right_index=True)
        summary = summary.merge(beta, how='left', left_index=True, right_index=True)
        summary = summary.merge(calendar_returns, how='left', left_index=True, right_index=True)
        summary = summary.astype(float).round(2)

        calendar_return_headings = calendar_returns.columns.to_frame().T

        return summary, calendar_return_headings

    def _get_arb_excess_return(self):
        fund_return = self._get_annualized_return()
        fund_return.rename(columns={"Return": "Fund"}, inplace=True)

        funds = self._updated_constituent_returns.columns
        benchmark_returns = self._absolute_benchmark_returns.reindex(funds, axis=1)

        benchmark_returns[self._updated_constituent_returns.isna()] = None

        bmrk_return = self._calculate_annualized_return(returns=benchmark_returns)
        bmrk_return.rename(columns={'Return': 'Bmrk'}, inplace=True)
        # bmrk_return = self._analytics.compute_trailing_return(
        #     ror=self._absolute_benchmark_returns,
        #     window=self._trailing_months,
        #     as_of_date=self._as_of_date,
        #     method="geometric",
        #     periodicity=Periodicity.Monthly,
        #     annualize=True,
        # )
        # bmrk_return = bmrk_return.to_frame('Bmrk')

        fund_bmrk_return = fund_return.merge(bmrk_return, left_index=True, right_index=True)
        fund_bmrk_return['Excess'] = fund_bmrk_return['Fund'] - fund_bmrk_return['Bmrk']
        excess_return = fund_bmrk_return['Excess'].to_frame('Excess')
        excess_return['Excess'] = excess_return['Excess'].astype(float).round(2)
        return excess_return

    def _get_arb_r_squared(self):
        funds = self._updated_constituent_returns.columns.tolist()
        r2_summary = pd.DataFrame(index=funds)
        for fund in funds:
            fund_return = self._updated_constituent_returns[fund].to_frame('Fund')
            fund_return = fund_return.astype(float)
            if fund in self._absolute_benchmark_returns.columns:
                bmrk_return = self._absolute_benchmark_returns[fund].to_frame('Bmrk')
                data = fund_return.merge(bmrk_return, left_index=True, right_index=True)
                correlation = data.corr(min_periods=24)
                r2 = correlation.loc['Fund', 'Bmrk'] ** 2
                r2 = r2.round(2)
                r2_summary.loc[fund, 'R2'] = r2
        return r2_summary

    def build_absolute_return_benchmark_summary(self):
        summary = self._get_arb_excess_return()
        # r_squared = self._get_arb_r_squared()
        # summary = excess_return.merge(r_squared, how='left', left_index=True, right_index=True)

        if summary.shape[1] == 0:
            return pd.DataFrame(columns=['Excess', 'ExcessDecile'], index=summary.index)

        summary['Excess'] = summary['Excess'] + np.random.random(summary.shape[0]) / 1e3
        summary['ExcessDecile'] = pd.qcut(x=summary['Excess'],
                                          q=[x / 100 for x in range(0, 110, 10)],
                                          labels=[x / 10 for x in range(10, 110, 10)][::-1])

        summary['ExcessDecile'] = summary['ExcessDecile'].astype(float)
        summary = summary[['Excess', 'ExcessDecile']]
        return summary

    def build_rba_excess_return_summary(self):
        inv_group_ids = self._constituents['InvestmentGroupId'].unique().tolist()
        inv_group = InvestmentGroup(investment_group_ids=inv_group_ids)
        rba = inv_group.get_rba_return_decomposition_by_date(
            start_date=self._start_date, end_date=self._end_date, frequency='M', window=36,
            factor_filter=['NON_FACTOR'])
        excess = self._calculate_annualized_return(returns=rba)
        excess = excess.reset_index()
        excess = excess[excess['AggLevel'] == 'NON_FACTOR']
        excess.columns = excess.columns[:-1].tolist() + ['Excess']
        excess.drop(columns={'InvestmentGroupId', 'FactorGroup1'}, inplace=True)
        excess = excess[excess['AggLevel'] == 'NON_FACTOR'].drop(columns='AggLevel')
        excess = excess.set_index('InvestmentGroupName')
        excess.rename(columns={'Excess': 'RbaAlpha'}, inplace=True)

        # need to ensure no points are identical. randomly break ties
        excess['RbaAlpha'] = excess['RbaAlpha'] + np.random.random(excess.shape[0]) / 1e3

        deciles = pd.qcut(x=excess['RbaAlpha'], q=[x / 100 for x in range(0, 110, 10)],
                          labels=[x / 10 for x in range(10, 110, 10)][::-1])
        deciles.name = 'RbaAlphaDecile'

        summary = excess.merge(deciles, left_index=True, right_index=True)
        summary['RbaAlphaDecile'] = summary['RbaAlphaDecile'].astype(float)

        return summary

    def build_rba_risk_decomposition_summary(self):
        inv_group_ids = self._constituents['InvestmentGroupId'].unique().tolist()
        inv_group = InvestmentGroup(investment_group_ids=inv_group_ids)
        rba_risk_decomp = inv_group.get_average_risk_decomp_by_group(
            start_date=self._start_date, end_date=self._end_date, frequency='M', window=36,
            group_type='FactorGroup1', wide=False)
        rba_risk_decomp.drop(columns='InvestmentGroupId', inplace=True)
        rba_risk_decomp = rba_risk_decomp.pivot(index='InvestmentGroupName', columns='FactorGroup1')
        rba_risk_decomp.columns = rba_risk_decomp.columns.droplevel(0)
        rba_risk_decomp = rba_risk_decomp.reindex(['SYSTEMATIC', 'X_ASSET_CLASS', 'PUBLIC_LS', 'NON_FACTOR'], axis=1)
        rba_risk_decomp = rba_risk_decomp.fillna(0)

        rba_risk_decomp['NON_FACTOR'] = rba_risk_decomp['NON_FACTOR'] + np.random.random(rba_risk_decomp.shape[0]) / 1e3
        rba_risk_decomp['PctIdioDecile'] = pd.qcut(x=rba_risk_decomp['NON_FACTOR'],
                                                   q=[x / 100 for x in range(0, 110, 10)],
                                                   labels=[x / 10 for x in range(10, 110, 10)][::-1]).astype(float)
        return rba_risk_decomp

    def build_summary_table(self, rankings, stats):
        quartile_headings = pd.DataFrame({'Rank': [-100] * 4,
                                       'Decile': [''] * 4,
                                       'Confidence': [''] * 4,
                                       'InvestmentGroupName': ['Quartile ' + str(x) for x in [1, 2, 3, 4]],
                                       'Quartile': [1, 2, 3, 4]})
        rankings = pd.concat([rankings, quartile_headings])

        summary = rankings.merge(stats, left_on=['InvestmentGroupName'], right_index=True, how='left')
        statuses = self._constituents[['InvestmentGroupName', 'InvestmentStatus']]
        summary = summary.merge(statuses, on='InvestmentGroupName', how='left')
        summary = summary.sort_values(['Quartile', 'Rank'])

        summary['Rank'] = [round(x, 0) if x != -100 else np.NAN for x in summary['Rank']]

        name_overrides = dict(zip(self._constituents['InvestmentGroupName'], self._constituents['InvestmentName']))
        summary["InvestmentGroupName"] = summary["InvestmentGroupName"].replace(name_overrides)
        summary["InvestmentGroupName"] = summary["InvestmentGroupName"].str.slice(0, 27)

        # test persistence over quarter ends
        as_of_dates = pd.date_range(dt.date(2019, 12, 31), self._as_of_date - relativedelta(days=1), freq='Q').tolist()
        as_of_dates = pd.to_datetime(as_of_dates).date.tolist()

        if len(as_of_dates) == 0:
            summary['Persistence'] = None
        else:
            current_rankings = summary[['InvestmentGroupName', 'Decile']]
            summary['Persistence'] = self._calculate_ranking_persistence(current_rankings=current_rankings,
                                                                         dates=as_of_dates)

        front_columns = ['Rank', 'Decile', 'Confidence', 'Persistence', 'InvestmentGroupName']

        col_order = front_columns + [x for x in summary.columns.tolist() if x not in set(front_columns + ['Quartile'])]
        summary = summary[col_order]
        return summary

    def build_omitted_fund_summary(self):
        constituents = self._constituents['InvestmentGroupName']
        modelled_constituents = self._updated_constituent_returns.columns.tolist()
        omitted_constituents = constituents[~constituents.isin(modelled_constituents)].dropna().tolist()
        omitted_constituents = sorted(omitted_constituents)

        returns = self._all_constituent_returns
        omitted_fund_returns = returns.loc[:, returns.columns.isin(omitted_constituents)]

        last_return_date = omitted_fund_returns.apply(lambda x: x.last_valid_index())
        last_return_date = last_return_date.to_frame('LastReturnDate')

        return_count = omitted_fund_returns.count(axis=0)
        return_count = return_count.to_frame('NoReturns')

        #TODO adjust
        partial_return = omitted_fund_returns.mean() * 12
        partial_return = partial_return.to_frame('PartialReturn')

        summary = pd.DataFrame(index=omitted_constituents)
        summary = summary.merge(last_return_date, left_index=True, right_index=True, how='left')
        summary = summary.merge(return_count, left_index=True, right_index=True, how='left')
        summary = summary.merge(partial_return, left_index=True, right_index=True, how='left')

        summary = summary.reset_index()

        return summary

    def _calculate_ranking_persistence(self, current_rankings, dates):
        rankings = current_rankings
        rankings.rename(columns={'Decile': self._as_of_date}, inplace=True)
        for date in dates:
            ranking = self._download_prior_rankings(location=self._summary_data_location,
                                                    as_of_date=date)
            if ranking is not None:
                ranking = ranking.drop_duplicates().groupby('InvestmentGroupName').head(1)
                rankings = rankings.merge(ranking, how='left')

        rankings.iloc[:, 1:] = rankings.iloc[:, 1:].transform(pd.to_numeric, errors='ignore')
        num_above = rankings.iloc[:, 1:].lt(6).sum(axis=1)
        num_obs = rankings.iloc[:, 1:].notna().sum(axis=1)
        current_above = rankings.iloc[:, 1].lt(6)
        pct_above = (num_above / num_obs).round(2)
        persistence = ([pct_above if current_above else 1 - pct_above for pct_above, num_obs, current_above
                        in zip(pct_above, num_obs, current_above)])
        return persistence

    def generate_performance_screener_report(self):
        # if not self._validate_inputs():
        #     return 'Invalid inputs'

        logging.info("Generating report for: " + self._peer_group)

        header_info = self.get_header_info()

        if self._updated_constituent_returns is None:
            return "No peers have returns through as of date"

        constituent_counts = self.build_constituent_count_summary()
        standalone_metrics, calendar_return_headings = self.build_standalone_metrics_summary()
        relative_metrics = self.build_absolute_return_benchmark_summary()
        rankings = self._calculate_peer_rankings()

        rba_excess_return_summary = self.build_rba_excess_return_summary()
        rba_risk_decomposition_summary = self.build_rba_risk_decomposition_summary()

        summary = standalone_metrics.merge(relative_metrics, how='left', left_index=True, right_index=True)
        summary = summary.merge(rba_excess_return_summary, how='left', left_index=True, right_index=True)
        summary = summary.merge(rba_risk_decomposition_summary, how='left', left_index=True, right_index=True)

        summary_table = self.build_summary_table(rankings=rankings, stats=summary)
        omitted_funds = self.build_omitted_fund_summary()

        rankings_max_row = 11 + summary_table.shape[0]
        rankings_max_column = get_column_letter(summary_table.shape[1] + 1)
        omitted_funds_max_row = 11 + omitted_funds.shape[0]
        omitted_funds_max_column = get_column_letter(omitted_funds.shape[1] + 1)
        print_areas = {'Rankings': 'B1:' + rankings_max_column + str(rankings_max_row),
                       'Omitted Funds': 'B1:' + omitted_funds_max_column + str(omitted_funds_max_row)}

        logging.info('Report summary data generated for: ' + self._peer_group)

        input_data = {
            "header_info1": header_info,
            "header_info2": header_info,
            "summary_table": summary_table,
            "calendar_return_headings": calendar_return_headings,
            "omitted_funds": omitted_funds,
            "constituents1": constituent_counts,
            "constituents2": constituent_counts,
        }

        input_data_json = {
            "header_info1": header_info.to_json(orient='index'),
            "header_info2": header_info.to_json(orient='index'),
            "summary_table": summary_table.to_json(orient='index'),
            "calendar_return_headings": calendar_return_headings.to_json(orient='index'),
            "omitted_funds": omitted_funds.to_json(orient='index'),
            "constituents1": constituent_counts.to_json(orient='index'),
        }

        data_to_write = json.dumps(input_data_json)
        write_params = AzureDataLakeDao.create_get_data_params(
            self._summary_data_location,
            self._get_json_file_name(as_of_date=self._as_of_date),
            retry=False,
        )
        self._runner.execute(
            params=write_params,
            source=DaoSource.DataLake,
            operation=lambda dao, params: dao.post_data(params, data_to_write),
        )

        logging.info("JSON stored to DataLake for: " + self._peer_group)

        as_of_date = dt.datetime.combine(self._as_of_date, dt.datetime.min.time())
        entity_name = self._peer_group.replace("/", "").replace("GCM ", "") + ' Peer'

        with Scenario(runner=DaoRunner(), as_of_date=as_of_date).context():
            InvestmentsReportRunner().execute(
                data=input_data,
                template="ARS_Performance_Screener_Template.xlsx",
                save=True,
                runner=self._runner,
                entity_type=ReportingEntityTypes.cross_entity,
                entity_name=entity_name,
                entity_display_name=entity_name,
                report_name="ARS Performance Screener",
                report_type=ReportType.Performance,
                report_vertical=ReportVertical.ARS,
                report_frequency="Monthly",
                aggregate_intervals=AggregateInterval.MTD,
                print_areas=print_areas,
                # output_dir="cleansed/investmentsreporting/printedexcels/",
                # report_output_source=DaoSource.DataLake,
            )

        logging.info("Excel stored to DataLake for: " + self._peer_group)

    def run(self, **kwargs):
        self.generate_performance_screener_report()
        return self._peer_group + " Complete"


if __name__ == "__main__":
    peer_groups = ["GCM Asia",
                   # "GCM Asia Credit",
                   "GCM Asia Equity",
                   # "GCM Asia Macro",
                   "GCM China",
                   # "GCM Commodities",
                   "GCM Consumer",
                   "GCM Credit",
                   "GCM Cross Cap",
                   "GCM Diverse",
                   "GCM Emerging Market Credit",
                   "GCM Energy",
                   "GCM Europe Credit",
                   "GCM Europe Equity",
                   "GCM Financials",
                   "GCM Fundamental Credit",
                   "GCM Generalist Long/Short Equity",
                   "GCM Healthcare",
                   "GCM Illiquid Credit",
                   "GCM Industrials",
                   "GCM Japan",
                   "GCM Long Only Equity",
                   "GCM Long/Short Credit",
                   "GCM Macro",
                   "GCM Merger Arbitrage",
                   "GCM Multi-PM",
                   "GCM Multi-Strategy",
                   "GCM Quant",
                   "GCM Real Estate",
                   "GCM Relative Value",
                   "GCM Short Sellers",
                   "GCM Structured Credit",
                   "GCM TMT",
                   "GCM Utilities",
                   ]

    peer_groups = ["GCM TMT"]

    runner = DaoRunner(
            container_lambda=lambda b, i: b.config.from_dict(i),
            config_params={
                DaoRunnerConfigArgs.dao_global_envs.name: {
                    DaoSource.InvestmentsDwh.name: {
                        "Environment": "prd",
                        "Subscription": "prd",
                    },
                    DaoSource.DataLake.name: {
                        "Environment": "prd",
                        "Subscription": "prd",
                    },
                }
            })

    as_of_dates = pd.date_range(dt.date(2019, 12, 31), dt.date(2022, 9, 30), freq='Q').tolist()
    as_of_dates = pd.to_datetime(as_of_dates).date.tolist()

    for peer_group in peer_groups:
        for as_of_date in as_of_dates:
            with Scenario(runner=runner, as_of_date=as_of_date).context():
                PerformanceScreenerReport(peer_group=peer_group).execute()
