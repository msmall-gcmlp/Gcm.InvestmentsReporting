from gcm.inv.scenario import Scenario, DaoRunner
from gcm.Dao.daos.azure_datalake.azure_datalake_dao import (
    AzureDataLakeFile,
    AzureDataLakeDao,
    TabularDataOutputTypes,
)
from openpyxl.worksheet.worksheet import Worksheet
from gcm.Dao.DaoRunner import DaoSource
from openpyxl import Workbook
from ..conversion_tools.excel_io import ExcelIO
from Reporting.core.components.report_table import ReportTable
from Reporting.core.components.report_workbook_handler import (
    ReportWorkBookHandler,
)
from ..conversion_tools.combine_excel import copy_sheet
from typing import List
from ..conversion_tools.convert_excel_to_pdf import convert
import io
import re
import pandas as pd
from Reporting.core.components.report_worksheet import ReportWorksheet


def get_template(
    blob_location: AzureDataLakeDao.BlobFileStructure,
) -> Workbook:
    assert blob_location is not None
    dao: DaoRunner = Scenario.get_attribute("dao")
    params = AzureDataLakeDao.create_blob_params(blob_location)
    file: AzureDataLakeFile = dao.execute(
        params=params,
        source=DaoSource.DataLake,
        operation=lambda d, p: d.get_data(p),
    )
    excel = file.to_tabular_data(
        TabularDataOutputTypes.ExcelWorkBook, params
    )
    return excel


def _trim_and_resize_rows(
    trim_rows: List[int], wb: Workbook, sheet: ReportWorksheet
) -> Workbook:
    for r in trim_rows:
        wb[sheet.worksheet_name].delete_rows(r)
        # have to reset row heights after deleting rows
        new_row_heights = _get_sheet_new_row_heights(
            wb[sheet.worksheet_name], trim_rows, 14.5
        )
        for i in range(0, len(new_row_heights)):
            # [i + 1] - because excel rows are numbered starting at 1
            wb[sheet.worksheet_name].row_dimensions[
                i + 1
            ].height = new_row_heights[i]
    return wb


def _generate_ws_level_trim_rows(
    wb: Workbook, sheet: ReportWorksheet
) -> List[int]:
    # trim_rows is a list of integers for excel rows to delete
    # blank rows within excel named ranges are deleted
    trim_rows = []
    for t in sheet.report_tables:
        if t.component_name in sheet.render_params.trim_region:
            for j in return_trim_rows(t, wb=wb):
                trim_rows.append(j)
                # rows have to be deleted in descending excel row order
    trim_rows = [x for x in set(trim_rows) if x is not None]
    trim_rows.sort(reverse=True)
    return trim_rows


def render_worksheet(wb: Workbook, sheet: ReportWorksheet):
    # to do: handle all the permutation combinations of a worksheet rendering
    # need to decide if we always need to wrap tables in a worksheet.
    assert sheet.worksheet_name is not None

    # first write all ReportTables for ReportingWorkSheet
    if sheet.report_tables is not None and len(sheet.report_tables) > 0:
        for t in sheet.report_tables:
            wb = print_table_component(wb, t)

    # next check format conditions
    if bool(sheet.render_params.trim_region):

        trim_rows = _generate_ws_level_trim_rows(wb, sheet)
        if len(trim_rows) > 0:
            wb = _trim_and_resize_rows(
                trim_rows=trim_rows, wb=wb, sheet=sheet
            )

    # hide_columns is list: ['A', 'B', 'C']
    if bool(sheet.render_params.hide_columns):
        for col in sheet.render_params.hide_columns:
            wb[sheet.worksheet_name].column_dimensions[col].hidden = True

    # print_region is string: "B1:AC150"
    if bool(sheet.render_params.print_region):
        wb[
            sheet.worksheet_name
        ].print_area = sheet.render_params.print_region

    return wb


def _get_sheet_new_row_heights(sheet, trim_rows, default_height=14.5):
    row_heights = pd.DataFrame()
    for i in range(0, sheet.max_row):
        row_heights = pd.concat(
            [
                row_heights,
                pd.DataFrame(
                    {
                        "row": [i + 1],
                        "height": [sheet.row_dimensions[i + 1].height],
                    }
                ),
            ]
        ).fillna(default_height)
    new_row_heights = list(
        row_heights[~row_heights.row.isin(trim_rows)].height
    )
    return new_row_heights


def return_trim_rows(k: ReportTable, wb: Workbook) -> List[int]:
    # the the range in question must be a rectangle
    # (i.e. someone can't just ctrl-clicked a bunch of random cells)
    address = list(wb.defined_names[k.component_name].destinations)
    excel_row_range = []
    rows_to_delete = []
    for sheetname, cell_address in address:
        for row_number in re.split("(\d+)", cell_address.replace("$", "")):
            try:
                excel_row_range.append(int(row_number))
            except ValueError:
                pass
    # can't trim if excel_row_range is not 2 integers
    rectangle_check = list(dict.fromkeys(excel_row_range))
    if len(rectangle_check) == 2:
        named_range_range = range(excel_row_range[0], excel_row_range[1])

        # no trimming to be done if length of dataframe is the same size as region
        if len(named_range_range) != len(k.df):
            rows_to_delete.extend(
                list(
                    range(
                        excel_row_range[0] + len(k.df),
                        excel_row_range[1] + 1,
                    )
                )
            )
    return rows_to_delete


def print_table_component(wb: Workbook, k: ReportTable) -> Workbook:
    if k.component_name in wb.defined_names:
        address = list(wb.defined_names[k.component_name].destinations)
        for sheetname, cell_address in address:
            if 'Table of Contents_' in sheetname:
                sheetname = 'Table of Contents'
            cell_address = cell_address.replace("$", "")
            # override wb:
            wb = ExcelIO.write_dataframe_to_xl(
                wb, k.df, sheetname, cell_address
            )
    return wb


def merge_files(wb_list: List[Workbook]):
    merged = wb_list[0]
    wb_count = 0
    for k in wb_list:
        if wb_count > 0:
            ws_count = 0
            for s in k.sheetnames:
                source_sheet: Worksheet = k[s]
                if 'Contents' in s:
                    continue

                target_sheet_name = f"{s}_{wb_count}_{ws_count}"
                merged.create_sheet(target_sheet_name)
                ws2: Worksheet = merged[target_sheet_name]
                copy_sheet(source_sheet, ws2)
                ws_count = ws_count + 1
        wb_count = wb_count + 1
    return merged


def generate_workbook(handler: ReportWorkBookHandler) -> Workbook:
    wb = get_template(handler.template_location)
    if handler.report_sheets is not None:
        for ws in handler.report_sheets:
            wb = render_worksheet(wb, ws)
    return wb


def print_pdf_report(
    params: dict, source=DaoSource, b: bytes = None, wb: Workbook = None
) -> dict:
    if b is None:
        if wb is not None:
            wb_stream = io.BytesIO()
            wb.save(wb_stream)
            b = wb_stream.getvalue()
        else:
            raise RuntimeError()
    save_params = convert(
        io.BytesIO(b),
        base_params=params,
        source=source,
    )
    return save_params


def print_excel_report(
    wb: Workbook,
    dao: DaoRunner,
    source: DaoSource,
    params: dict,
    save: bool,
    print_pdf: bool = True,
) -> dict:
    if save:
        wb_stream = io.BytesIO()
        wb.save(wb_stream)
        b = wb_stream.getvalue()
        dao.execute(
            params=params,
            source=source,
            operation=lambda d, v: d.post_data(v, b),
        )
        if print_pdf:
            print_pdf_report(params, source, b)

    return params
