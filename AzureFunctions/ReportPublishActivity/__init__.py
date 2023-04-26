from gcm.inv.utils.azure.durable_functions.base_activity import (
    BaseActivity,
)
from ..utils.reporting_parsed_args import ReportingParsedArgs, ReportNames
import json
from gcm.inv.scenario import Scenario, DaoRunner
from gcm.Dao.daos.azure_datalake.azure_datalake_dao import (
    AzureDataLakeFile,
    AzureDataLakeDao,
    TabularDataOutputTypes,
)
from openpyxl.worksheet.worksheet import Worksheet
from gcm.Dao.DaoRunner import DaoSource
from ..Reporting.core.report_structure import ReportStructure
from ..Reporting.Reports.controller import get_report_class_by_name
from azure.core.exceptions import ResourceNotFoundError
from openpyxl import Workbook
from ..utils.conversion_tools.excel_io import ExcelIO
from ..Reporting.core.components.report_table import ReportTable
from ..Reporting.core.components.report_workbook_handler import (
    ReportWorkBookHandler,
)
from ..utils.conversion_tools.convert_excel_to_pdf import convert
from ..utils.conversion_tools.combine_excel import copy_sheet
import io
from typing import List


class ReportPublishActivity(BaseActivity):
    # TODO: Decide if these should be in RAW

    def __init__(self):
        super().__init__()
        self.excel_template_cache = {}

    @property
    def parg_type(self):
        return ReportingParsedArgs

    def get_template(
        self, blob_location: AzureDataLakeDao.BlobFileStructure
    ):
        assert blob_location is not None
        dao: DaoRunner = Scenario.get_attribute("dao")
        try:
            params = AzureDataLakeDao.create_blob_params(blob_location)
            file: AzureDataLakeFile = dao.execute(
                params=params,
                source=DaoSource.DataLake,
                operation=lambda d, p: d.get_data(p),
            )
            excel = file.to_tabular_data(
                TabularDataOutputTypes.ExcelWorkBook, params
            )
            return excel
        except ResourceNotFoundError:
            return None

    @staticmethod
    def print_report_to_template(wb: Workbook, struct: ReportStructure):
        ReportPublishActivity.print_tables_to_range(
            wb, [x for x in struct.components if type(x) == ReportTable]
        )

    @staticmethod
    def print_tables_to_range(wb: Workbook, components: List[ReportTable]):
        for k in components:
            if k.component_name in wb.defined_names:
                address = list(
                    wb.defined_names[k.component_name].destinations
                )
                for sheetname, cell_address in address:
                    cell_address = cell_address.replace("$", "")
                    # override wb:
                    wb = ExcelIO.write_dataframe_to_xl(
                        wb, k.df, sheetname, cell_address
                    )
        return wb

    def activity(self, **kwargs):
        data = json.loads(kwargs["context"])["d"]["data"]
        params_vals = []
        # TODO: figure out if can parallelize the excel stuff
        for i in data:
            # get data from dao:
            dao: DaoRunner = Scenario.get_attribute("dao")
            file: AzureDataLakeFile = dao.execute(
                params=json.loads(i),
                source=DaoSource.DataLake,
                operation=lambda d, p: d.get_data(p),
            )
            assert file is not None
            d = json.loads(file.content)
            r = ReportNames[d["report_name"]]
            report_structure_class = get_report_class_by_name(r)
            report_structure: ReportStructure = (
                report_structure_class.from_dict(d, report_name=r)
            )
            assert report_structure is not None
            [params, source] = report_structure.save_params()

            # TODO: Not sure why this happens with ReportingStorage
            f_s = "filesystem_name"
            if (
                source == DaoSource.ReportingStorage
                and type(params) == dict
                and f_s in params
            ):
                params[f_s] = params[f_s].strip("/")
            if all(
                [
                    type(x) == ReportTable
                    for x in report_structure.components
                ]
            ):
                template = self.get_template(
                    report_structure.excel_template_location
                )
                if template is None:
                    # no template, get standard template
                    raise NotImplementedError(
                        "Must Have a Template (for now!)"
                    )
                else:
                    template: Workbook = template
                    final_template = (
                        ReportPublishActivity.print_report_to_template(
                            template, report_structure
                        )
                    )
                    wb_stream = io.BytesIO()
                    final_template.save(wb_stream)
                    b = wb_stream.getvalue()
                    if self.pargs.save:
                        dao.execute(
                            params=params,
                            source=source,
                            operation=lambda d, v: d.post_data(v, b),
                        )
                        convert(
                            io.BytesIO(b),
                            base_params=params,
                            source=source,
                        )

            else:
                wb_list: List[Workbook] = []
                for i in report_structure.components:
                    if type(i) == ReportWorkBookHandler:
                        handler: ReportWorkBookHandler = i
                        template = self.get_template(
                            handler.template_location
                        )
                        printed = self.print_tables_to_range(
                            template, handler.report_tables
                        )
                        wb_list.append(printed)
                assert len(wb_list) > 0
                merged = ReportPublishActivity.merge_files(wb_list)
                wb_stream = io.BytesIO()
                merged.save(wb_stream)
                b = wb_stream.getvalue()
                if self.pargs.save:
                    dao.execute(
                        params=params,
                        source=source,
                        operation=lambda d, v: d.post_data(v, b),
                    )
                    convert(
                        io.BytesIO(b),
                        base_params=params,
                        source=source,
                    )
            params = {
                key: value
                for key, value in params.items()
                if key
                in [
                    "filesystem_name",
                    "file_path",
                    "retry",
                    "metadata",
                ]
            }
            params_vals.append(params)
        return json.dumps(params_vals)

    @staticmethod
    def merge_files(wb_list: List[Workbook]):
        merged = wb_list[0]
        wb_count = 0
        for k in wb_list:
            if wb_count > 0:
                ws_count = 0
                for s in k.sheetnames:
                    source_sheet: Worksheet = k[s]
                    target_sheet_name = f"{s}_{wb_count}_{ws_count}"
                    merged.create_sheet(target_sheet_name)
                    ws2: Worksheet = merged[target_sheet_name]
                    copy_sheet(source_sheet, ws2)
                    # for row in merge_sheet.iter_rows():
                    #     for cell in row:
                    #         col_letter = get_column_letter(cell.column)
                    #         cell_coord = f"{col_letter}{cell.row}"
                    #         cell_value = cell.value
                    #         ws2[cell_coord].value = cell_value
                    #         ws2[cell_coord]._style = cell._style
                    ws_count = ws_count + 1
            wb_count = wb_count + 1
        return merged


def main(context):
    return ReportPublishActivity().execute(context=context)
