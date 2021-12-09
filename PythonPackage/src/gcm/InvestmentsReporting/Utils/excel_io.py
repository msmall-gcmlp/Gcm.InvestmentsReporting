import re
import numpy as np
from openpyxl.styles import Alignment, Font

from openpyxl.utils.cell import column_index_from_string
from openpyxl.utils.dataframe import dataframe_to_rows


class ExcelIO(object):
    def write_dataframe_to_xl(
        self,
        wb,
        dataframe,
        template_sheet_name,
        first_cell,
        alignment=Alignment(),
        font=Font(),
    ):
        ws = wb[template_sheet_name]
        first_cell = re.split("(\d+)", first_cell)
        start_row = first_cell[1]
        start_col = first_cell[0]

        numerics = dataframe.select_dtypes(include=[np.number])
        dataframe[numerics.columns] = np.round(
            dataframe[numerics.columns], 6
        )

        rows = dataframe_to_rows(dataframe, index=False, header=False)
        for r_idx, row in enumerate(rows, 0):
            for c_idx, value in enumerate(row, 0):
                ws.cell(
                    row=int(start_row) + r_idx,
                    column=column_index_from_string(start_col) + c_idx,
                    value=value,
                )
                ws.cell(
                    row=int(start_row) + r_idx,
                    column=column_index_from_string(start_col) + c_idx,
                ).alignment = alignment
                ws.cell(
                    row=int(start_row) + r_idx,
                    column=column_index_from_string(start_col) + c_idx,
                ).font = font
        return wb

    def write_series_to_xl_long(
        self, wb, series, template_sheet_name, first_cell
    ):
        # TODO refactor name to xl_long
        ws = wb[template_sheet_name]
        first_cell = re.split("(\d+)", first_cell)
        start_row = first_cell[1]
        column_letter = first_cell[0]
        column_index = column_index_from_string(column_letter)

        series = series.round(6) if series.dtype == "float64" else series

        for r_idx, value in enumerate(series):
            ws.cell(
                row=int(start_row) + r_idx,
                column=column_index,
                value=value,
            )

        return wb

    def write_series_to_xl_wide(
        self, wb, series, template_sheet_name, first_cell
    ):
        ws = wb[template_sheet_name]
        first_cell = re.split("(\d+)", first_cell)
        row_index = first_cell[1]
        column_letter = first_cell[0]
        start_col = column_index_from_string(column_letter)

        series = series.round(6) if series.dtype == "float64" else series

        for c_idx, value in enumerate(series):
            ws.cell(
                row=int(row_index), column=start_col + c_idx, value=value
            )

        return wb
