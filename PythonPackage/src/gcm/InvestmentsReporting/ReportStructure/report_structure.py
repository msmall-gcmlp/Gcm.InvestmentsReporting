from abc import ABC
from enum import Enum
import logging
from gcm.Scenario.scenario_enums import AggregateInterval
from gcm.Scenario.scenario import Scenario
from gcm.Dao.daos.azure_datalake.azure_datalake_dao import AzureDataLakeDao
from gcm.Dao.daos.azure_datalake.azure_datalake_file import (
    AzureDataLakeFile,
)
from gcm.Dao.Utils.tabular_data_util_outputs import TabularDataOutputTypes
from gcm.Dao.DaoRunner import DaoRunner
from gcm.Dao.DaoSources import DaoSource
import openpyxl
from ..Utils.excel_io import ExcelIO
from openpyxl.writer.excel import save_virtual_workbook
import datetime as dt
import json
from openpyxl import Workbook

template_location = (
    "/".join(
        [
            "raw",
            "test",
            "rqstest",
            "rqstest",
            "ReportingTemplates",
        ]
    )
    + "/"
)

base_output_location = (
    "/".join(["lab", "rqs", "Reports", "Scripted Outputs"]) + "/"
)


class ReportStage(Enum):
    PreDeployment = (0,)
    IC = (1,)
    Active = (2,)
    Legacy = (3,)


class ReportVertical(Enum):
    PEREI = (0,)
    ARS = (1,)
    SIG = (2,)
    FirmWide = (3,)


class ReportType(Enum):
    Risk = (0,)
    CapitalAndExposure = (1,)
    Performance = (2,)
    CommitmentsAndFlows = (3,)


class ReportSubstrategy(Enum):
    Credit = (0,)
    Equities = (1,)
    Primaries = (2,)
    Secondaries = (3,)
    Infrastructure = (4,)
    All = (5,)


class RiskReportConsumer(Enum):
    RiskMonitoring = (0,)
    InternalExRMA = (1,)
    CIO = (2,)
    External = (3,)


class ReportingEntityTypes(Enum):
    portfolio = (0,)
    manager_fund = (1,)
    cross_entity = (2,)


class ReportingEntityTag(object):
    def __init__(
        self,
        entity_type: ReportingEntityTypes,
        entity_name: str,
        display_name: str,
        entity_id: int,
        source: DaoSource,
        runner: DaoRunner,
    ):
        self.entity_type = entity_type
        self.entity_name = entity_name
        self.display_name = display_name
        self._entity_id_holder = entity_id
        self._entity_id = None
        self.entity_source = source
        self.runner = runner

    def to_metadata_tags(self):
        if self.get_entity_id() is not None:
            return {
                f"gcm_{self.entity_type.name}_id": str(
                    self.get_entity_id()
                )
            }
        return None

    def get_entity_id(self):
        if self._entity_id is None:
            if self._entity_id_holder is not None:
                # simply set and forget.
                self._entity_id = self._entity_id_holder
            elif self.entity_source is not None:
                # utilize the DAO to attempt to get an ID
                self._entity_id = self._generate_entity_id()
        return self._entity_id

    def _generate_entity_id(self):
        # TODO: This is TEMPORARY
        # replace logic with EntityHierarchy
        params = {}
        get_id_by = ""
        if self.entity_source == DaoSource.PubDwh:
            if self.entity_type == ReportingEntityTypes.portfolio:
                params = {
                    "table_name": "Portfolios",
                    "schema": "Analytics",
                    # eventually paramterize this further...
                    "operation": lambda query, item: query.filter(
                        item.Acronym == self.entity_name
                    ),
                }
                get_id_by = "PortfolioMasterId"
            if self.entity_type == ReportingEntityTypes.manager_fund:
                s: Scenario = Scenario.current_scenario()
                params = {
                    "table_name": "PortfolioInvestmentBalances",
                    "schema": "Analytics",
                    # eventually paramterize this further...
                    "operation": lambda query, item: query.filter(
                        item.InvestmentName == self.entity_name,
                        item.PeriodDate == s.get_attribute("asofdate"),
                    ),
                }
                get_id_by = "InvestmentMasterId"

        data = self.runner.execute(
            params=params,
            source=self.entity_source,
            operation=lambda dao, params: dao.get_data(params),
        )
        selected = data[get_id_by]
        value = list(selected)[0]
        return value


# seperate class in case we want to load once
# and pass to multiple report structures
class ReportTemplate(object):
    def __init__(
        self, filename, runner, template_location=template_location
    ):
        self.filename = filename
        self._excel = None
        self.runner: DaoRunner = runner
        self.template_location = template_location

    def excel(self, excel_params: dict = {}) -> openpyxl.Workbook:
        if self._excel is None:
            # standard location

            params = AzureDataLakeDao.create_get_data_params(
                self.template_location,
                self.filename,
                retry=False,
            )
            file: AzureDataLakeFile = self.runner.execute(
                params=params,
                source=DaoSource.DataLake,
                operation=lambda dao, params: dao.get_data(params),
            )
            for k in excel_params:
                params[k] = excel_params[k]
            self._excel = file.to_tabular_data(
                output_type=TabularDataOutputTypes.ExcelWorkBook,
                params=params,
            )
        return self._excel


# https://gcmlp1.atlassian.net/
# wiki/spaces/IN/pages/2719186981/
# Metadata+Fields
class ReportStructure(ABC):
    def __init__(
        self,
        report_name,
        data,
        asofdate,
        runner,
        aggregate_intervals=[AggregateInterval.Daily],
        report_types=[ReportType.Risk],
        stage=ReportStage.Active,
        report_vertical=[ReportVertical.FirmWide],
        report_substrategy=[ReportSubstrategy.All],
        report_consumers=[RiskReportConsumer.RiskMonitoring],
    ):
        self.report_name = report_name
        self.data = data

        # gcm tags
        self.gcm_as_of_date = asofdate
        self.gcm_report_period = aggregate_intervals
        self.gcm_report_type = report_types
        self.gcm_report_target_stage = stage
        self.gcm_business_group = report_vertical
        self.gcm_strategy = report_substrategy
        self.gcm_target_audience = report_consumers

        self.template: ReportTemplate = None
        self._workbook: openpyxl.Workbook = None
        self._report_entity: ReportingEntityTag = None
        self._runner = runner

    def load_template(self, template: ReportTemplate):
        if self.template is None:
            self.template = template
        else:
            logging.log("Template info has already been set")

    def load_workbook(self, workbook: openpyxl.Workbook):
        if self._workbook is None:
            self._workbook = workbook
        else:
            logging.log("Wb has already been set")

    def load_reporting_entity(self, entity_tag: ReportingEntityTag):
        if self._report_entity is None:
            self._report_entity = entity_tag
        else:
            logging.log("Entity has already been set")

    def print_report(self, **kwargs):
        output_dir = kwargs.get("output_dir", base_output_location)
        excel_io = ExcelIO()
        wb: openpyxl.Workbook = None
        if self.template is not None:
            wb: openpyxl.Workbook = self.template.excel()
            print("going to attempt to print to template")
            for k in self.data:
                address = list(wb.defined_names[k].destinations)
                for sheetname, cell_address in address:
                    cell_address = cell_address.replace("$", "")
                    # override wb:
                    wb = excel_io.write_dataframe_to_xl(
                        wb, self.data[k], sheetname, cell_address
                    )
        elif self._workbook is not None:
            # we are in the case where
            #   data is present already
            #   all formatting is already done
            # and simply want to render report
            # using report structure.

            wb = self._workbook
        else:
            wb: openpyxl.Workbook = Workbook()
            excel_io = ExcelIO()
            for k in self.data:
                wb.create_sheet(title=k)
                wb = excel_io.write_dataframe_to_xl(
                    wb, self.data[k], k, cell_address
                )
        params = AzureDataLakeDao.create_get_data_params(
            output_dir,
            self.output_name(),
            metadata=self.serialize_metadata(),
        )
        b = save_virtual_workbook(wb)
        if kwargs.get("save", False):
            self._runner.execute(
                params=params,
                source=DaoSource.DataLake,
                operation=lambda d, v: d.post_data(v, b),
            )

    def output_name(self):
        s = ""
        if self._report_entity is not None:
            s += f"{self._report_entity.display_name}_"
        s += f"{self.report_name}_"
        s += f'{self.gcm_as_of_date.strftime("%Y-%m-%d")}.xlsx'
        return s

    def serialize_metadata(self):
        # convert tags from above
        # to json-serializable dictionary
        d = {}
        all_data = self.__dict__
        for k in all_data:
            k: str = k
            if k.startswith("gcm_"):
                # we want to serialize this:
                val = all_data[k]
                metadata = None
                if type(val) == dt.datetime:
                    metadata = val.strftime("%Y-%m-%d")
                elif type(val) == list:
                    if len(val) > 0:
                        if all(issubclass(type(f), Enum) for f in val):
                            metadata = json.dumps(
                                list(map(lambda x: x.name, val))
                            )

                elif issubclass(type(val), Enum):
                    metadata = val.name
                elif type(val) == str:
                    metadata = val
                if metadata is not None:
                    d[k] = metadata
        if self._report_entity is not None:
            # merge
            d2 = self._report_entity.to_metadata_tags()
            d.update(d2)
        return d
