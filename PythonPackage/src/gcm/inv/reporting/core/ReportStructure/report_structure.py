from abc import ABC
from ..Utils.convert_excel_to_pdf import convert
from gcm.inv.utils.misc.extended_enum import Enum, ExtendedEnum
import logging
from gcm.Scenario.scenario import AggregateInterval
from gcm.Dao.daos.azure_datalake.azure_datalake_dao import AzureDataLakeDao
from gcm.Dao.daos.azure_datalake.azure_datalake_file import (
    AzureDataLakeFile,
)
from gcm.Dao.Utils.tabular_data_util_outputs import TabularDataOutputTypes
from gcm.Dao.DaoRunner import DaoRunner
from gcm.Dao.DaoSources import DaoSource
import openpyxl
from ..Utils.excel_io import ExcelIO
from openpyxl.writer.excel import save_virtual_workbook
import datetime as dt
import json
from openpyxl import Workbook
import io
from typing import List


template_location = (
    "/".join(
        [
            "raw",
            "investmentsreporting",
            "exceltemplates",
        ]
    )
    + "/"
)

base_output_location = "performance/"
# base_output_location = "cleansed/investmentsreporting/printedexcels/"
daily = AggregateInterval.Daily


class ReportType(ExtendedEnum):
    Risk = "Risk"
    Capital_and_Exposure = "Capital & Exposure"
    Performance = "Performance"
    Commitments_and_Flows = "Commitments & Flows"
    Market = "Market"


class ReportStage(Enum):
    PreDeployment = 0
    IC = 1
    Active = 2
    Legacy = 3


class ReportVertical(Enum):
    PEREI = 0
    ARS = 1
    SIG = 2
    FirmWide = 3


class ReportStrategy(Enum):
    Credit = 0
    Equities = 1
    Primaries = 2
    Secondaries = 3
    Infrastructure = 4
    All = 5


class RiskReportConsumer(Enum):
    Risk = 0
    ARS_IC = 1
    ARS_PM = 2
    ARS_Research = 3
    PEREI_Deal_Team = 4
    PEREI_PM = 5
    PEREI_IC = 6
    CIO = 7


class ReportingEntityTypes(Enum):
    portfolio = 0
    manager_fund = 1  # investment
    manager_fund_group = 2  # investmentgroup
    cross_entity = 3


class ReportingEntityTag(object):
    def __init__(
        self,
        entity_type: ReportingEntityTypes,
        entity_name: str,
        display_name: str,
        entity_ids: List[int],
    ):
        self.entity_type = entity_type
        self.entity_name = entity_name
        self.display_name = display_name
        self._entity_id_holder = entity_ids
        self._entity_id = None

    def to_metadata_tags(self):
        if self.get_entity_ids() is not None:
            list_of_ids = self.get_entity_ids()
            if list_of_ids is not None:
                metadata = json.dumps(list(map(lambda x: x, list_of_ids)))
                return {f"gcm_{self.entity_type.name}_ids": metadata}
        return None

    def get_entity_ids(self):
        if self._entity_id is None:
            if self.entity_type != ReportingEntityTypes.cross_entity:
                if self._entity_id_holder is not None:
                    # simply set and forget.
                    self._entity_id = self._entity_id_holder
                else:
                    raise NotImplementedError("You must have passed in Entity Id Holder")
        return self._entity_id


# seperate class in case we want to load once
# and pass to multiple report structures
class ReportTemplate(object):
    def __init__(self, filename, runner, template_location=template_location):
        self.filename = filename
        self._excel = None
        self.runner: DaoRunner = runner
        self.template_location = template_location

    def excel(self, excel_params: dict = {}) -> openpyxl.Workbook:
        if self._excel is None:
            # standard location

            params = AzureDataLakeDao.create_get_data_params(
                self.template_location,
                self.filename,
                retry=False,
            )
            file: AzureDataLakeFile = self.runner.execute(
                params=params,
                source=DaoSource.DataLake,
                operation=lambda dao, params: dao.get_data(params),
            )
            for k in excel_params:
                params[k] = excel_params[k]
            self._excel = file.to_tabular_data(
                output_type=TabularDataOutputTypes.ExcelWorkBook,
                params=params,
            )
        return self._excel


# https://gcmlp1.atlassian.net/
# wiki/spaces/IN/pages/2719186981/
# Metadata+Fields
class ReportStructure(ABC):
    _display_mapping_dict = {
        ReportingEntityTypes.manager_fund: "PFUND",
        ReportingEntityTypes.manager_fund_group: "PFUND",
        ReportingEntityTypes.portfolio: "PORTFOLIO",
        ReportingEntityTypes.cross_entity: "XENTITY",
    }

    def __init__(
        self,
        report_name,
        data,
        asofdate: dt.datetime,
        runner: DaoRunner,
        report_type=ReportType.Risk,
        report_frequency="Monthly",
        aggregate_intervals=None,
        stage=None,
        report_vertical=ReportVertical.ARS,
        report_substrategy=None,
        report_consumers=None,
        **kwargs,
    ):
        self.data = data

        # Standards
        self.gcm_report_name = report_name
        self.gcm_report_type = report_type

        # gcm tags
        self.gcm_as_of_date = asofdate
        self.gcm_report_period = aggregate_intervals
        self.gcm_report_target_stage = stage
        self.gcm_business_group = report_vertical
        self.gcm_strategy = report_substrategy
        self.gcm_target_audience = report_consumers
        self.gcm_modified_date = dt.datetime.utcnow().strftime("%Y-%m-%d")
        self.gcm_report_frequency = report_frequency
        self.template: ReportTemplate = None
        self._workbook: openpyxl.Workbook = None
        self._report_entity: ReportingEntityTag = None
        self._raw_pdf = None
        self._runner = runner

    def load_template(self, template: ReportTemplate):
        if self.template is None:
            self.template = template
        else:
            logging.log("Template info has already been set")

    def load_workbook(self, workbook: openpyxl.Workbook):
        if self._workbook is None:
            self._workbook = workbook
        else:
            logging.log("Wb has already been set")

    def load_reporting_entity(self, entity_tag: ReportingEntityTag):
        if self._report_entity is None:
            self._report_entity = entity_tag
        else:
            logging.log("Entity has already been set")

    def load_pdf(self, location, file_name):
        if self._raw_pdf is None:
            params = AzureDataLakeDao.create_get_data_params(
                location,
                file_name,
                retry=False,
            )
            self._raw_pdf = self._runner.execute(
                params=params,
                source=DaoSource.DataLake,
                operation=lambda dao, params: dao.get_data(params),
            )
        else:
            logging.log("pdf has already been set")

    def print_report(self, **kwargs):
        output_dir = (
            kwargs.get(
                "output_dir",
                {base_output_location},
            )
        ) + self.gcm_report_type.name

        if self._raw_pdf is not None:
            b = self._raw_pdf.content
        else:
            excel_io = ExcelIO()
            wb: openpyxl.Workbook = None
            if self.template is not None:
                wb: openpyxl.Workbook = self.template.excel()
                print("going to attempt to print to template")
                for k in self.data:
                    address = list(wb.defined_names[k].destinations)
                    for sheetname, cell_address in address:
                        cell_address = cell_address.replace("$", "")
                        # override wb:
                        wb = excel_io.write_dataframe_to_xl(wb, self.data[k], sheetname, cell_address)
                print_areas = kwargs.get("print_areas", None)
                if print_areas is not None:
                    for k, v in print_areas.items():
                        wb[k].print_area = v
            elif self._workbook is not None:
                # we are in the case where
                #   data is present already
                #   all formatting is already done
                # and simply want to render report
                # using report structure.

                wb = self._workbook
            else:
                wb: openpyxl.Workbook = Workbook()
                excel_io = ExcelIO()
                for k in self.data:
                    wb.create_sheet(title=k)
                    wb = excel_io.write_dataframe_to_xl(wb, self.data[k], k, cell_address)
            b = save_virtual_workbook(wb)
        params = AzureDataLakeDao.create_get_data_params(
            output_dir,
            self.output_name(),
            metadata=self.serialize_metadata(),
        )

        if kwargs.get("save", False):
            self._runner.execute(
                params=params,
                source=kwargs.get("output_source", DaoSource.ReportingStorage),
                operation=lambda d, v: d.post_data(v, b),
            )
        if kwargs.get("save_as_pdf", True):
            convert(
                io.BytesIO(b),
                self._runner,
                kwargs.get("output_source", DaoSource.ReportingStorage),
                params,
            )

    def output_name(self):
        s = f"{self.gcm_report_name}_"
        if self._report_entity is not None:
            s += f"{self._report_entity.display_name}_"
            entity_type_display = ReportStructure._display_mapping_dict[self._report_entity.entity_type]
            s += f"{entity_type_display}_"
        s += f"{self.gcm_report_type.name}_"
        s += f'{self.gcm_as_of_date.strftime("%Y-%m-%d")}.xlsx'
        return s

    def _get_metadata_from_property(self, all_data, k):
        metadata = None
        if k.startswith("gcm_"):
            # we want to serialize this:
            val = all_data[k]
            metadata = None
            if type(val) == dt.datetime:
                metadata = val.strftime("%Y-%m-%d")
            elif type(val) == list:
                if len(val) > 0:
                    if all(issubclass(type(f), Enum) for f in val):
                        metadata = json.dumps(list(map(lambda x: x.name, val)))
                    else:
                        metadata = json.dumps(list(map(lambda x: x, val)))
            elif issubclass(type(val), ExtendedEnum):
                metadata = val.value
            elif issubclass(type(val), Enum):
                metadata = val.name
            elif type(val) == str:
                metadata = val
        return metadata

    def serialize_metadata(self):
        # convert tags from above
        # to json-serializable dictionary
        d = {}
        all_data = self.__dict__
        for k in all_data:
            k: str = k
            metadata = self._get_metadata_from_property(all_data, k)
            if metadata is not None:
                d[k] = metadata
        if self._report_entity is not None:
            # merge
            d2 = self._report_entity.to_metadata_tags()
            if d2 is not None:
                d.update(d2)
        return d
